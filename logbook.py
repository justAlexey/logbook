import openpyxl as opxl
from progress.bar import IncrementalBar
import datetime
import database

class Logbook:

    def __init__(self):
        # try to open 'APN'.xls
        try:
            self.wb1225 = opxl.open("excel/1225.xlsx")
        except:
            print("Не могу открыть таблицу 1225.xlsx")
            exit()

        try:
            self.wb1328 = opxl.open("excel/1328.xlsx")
        except:
            print("Не могу открыть таблицу 1328.xlsx")
            exit()

        try:
            self.wb1498 = opxl.open("excel/1498.xlsx")
        except:
            print("Не могу открыть таблицу 1498.xlsx")
            exit()

        try:
            self.database = opxl.open("excel/database.xlsx")
        except:
            print("Не могу открыть базу")
            exit()

        # load worksheets
        self.ws1225 = self.wb1225.active
        self.ws1328 = self.wb1328.active
        self.ws1498 = self.wb1498.active
        self.wsDatabase = self.database.active
        self.wblb = opxl.Workbook("Logbook.xlsx")
        self.wslb = self.wblb.active

        # get nessesary columns` number 
        self.tow = self.__getInfo(self.ws1225, "ToW")
        self.cmproject = self.__getInfo(self.ws1225, "CM Project")
        self.sbt = self.__getInfo(self.ws1225, "User Sign")
        self.date = self.__getInfo(self.ws1225, "End Date")
        self.description = self.__getInfo(self.ws1328, "Description")
        self.pn = self.__getInfo(self.ws1328, "Part No.")
        self.cmproject1328 = self.__getInfo(self.ws1328, "CM Project No.")
        self.sn = self.__getInfo(self.ws1328, "Serial No. / Batch No.")
        self.timeDuration = self.__getInfo(self.ws1225, "Used MHR")
        self.ata = self.__getInfo(self.ws1328, "ATA Chapter")
        self.crs = self.__getInfo(self.ws1498, "Insp. by")
        self.serial1328 = self.__getInfo(self.ws1328, "Serial No. / Batch No.")
        self.serial1498 = self.__getInfo(self.ws1498, "Serial/Batch No.")
        self.part1328 = self.__getInfo(self.ws1328, "Part No.")
        self.part1498 = self.__getInfo(self.ws1498, "Part No.")
        self.date1498 = self.__getInfo(self.ws1498, "Insp. Date")


    def __getInfo(self, ws, info): # find column`s number in worksheet by info
        for i in range(ws.max_column):
            if ws[1][i].value == info:
                return i


    def clear1328(self, ws):    # clear worksheet from empty rows by empty cm project
        clearing = IncrementalBar("Clearing 1328...", max = ws.max_row-1)
        for i in range(ws.max_row,1,-1):
            if ws[i][self.cmproject1328].value == None:
                ws.delete_rows(i)
            clearing.next()
        clearing.finish()
        return ws


    def clear1225(self, ws):    # clear worksheet from empty rows by type of work
        clearing = IncrementalBar("Clearing 1225...", max = ws.max_row-1)
        for i in range(ws.max_row,1,-1):
            if ws[i][self.tow].value not in database.typeofwork:
                ws.delete_rows(i)
            clearing.next()
        clearing.finish()
        return ws


    def formateRow(self, row):      # formate logbook row like a array of string
        row1225 = self.ws1225[row]
        row1328 = self.getRow1328(row1225[self.cmproject])
        usedTime = self.getTime(row1225[self.timeDuration])
        task = self.getTaskType(row1225[self.tow].value)
        crs = self.getCRS(row1225, row1328)
        part = row1328[self.part1328].value
        serial = row1328[self.serial1328].value
        data = [
                row1225[self.date].value.strftime('%d.%m.%y'),    # date
                "ACCMS",    # location "ACCMS"
                self.getDescription(part),    # Component type (description)
                serial,    # Serial number
                "C6 Equipment",    # Type of maintenance(raiting)
                "trainee",    # Privelege
                task["fot"],    # task type: FOT 
                task["sgh"],    # task type: SGH
                task["r/i"],    # task type: R/I
                task["ts"],    # task type: TS
                task["mod"],    # task type: MOD
                task["rep"],    # task type: REP
                task["insp"],    # task type: INSP
                "",    # type of activity: Training
                "X",    # type of activity: Perform
                "",    # type of activity: Suppervise
                crs,    # type of activity: CRS
                self.getATA(part),    # ATA
                self.getOperation(crs),    # Operation performed
                usedTime,    # Time duration
                row1225[self.cmproject].value,    # Maintenance record ref.(CM number)
                "ToW is \"{}\"".format(row1225[self.tow].value)    # Remarks
        ]
        return data


 
    def writeRow(self, sheet, data):    # write row in the sheet with SBT name
        if sheet not in self.wblb.sheetnames:
            ws = self.wblb.create_sheet(sheet)
            ws.title = sheet
        else:
            ws = self.wblb[sheet]
        #print("sheet is {}, data is {}".format(sheet,data))
        ws.append(data)


    def getRow1328(self, cm):       # get row from 1328 apn by CM project
        low = 0
        high = self.ws1328.max_row
        while low < high:
            midle = (low+high)/2
            midle =  int(midle)
            row1328 = self.ws1328[midle]
            #print(f'low = {low}, high = {high}, max row = {self.ws1328.max_row}midle = {midle}, cm = {cm.value}, row cm = {row1328[self.cmproject1328].value}')
            if cm.value == row1328[self.cmproject1328].value:
                return row1328
            elif cm.value > row1328[self.cmproject1328].value:
                low = midle+1
            elif cm.value < row1328[self.cmproject1328].value:
                high = midle
        raise Exception('not found cm project in 1328')
        return -1

    def getTime(self, time):        # get time by delta seconds
        hours, remainder = divmod(time.value.seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours}:{minutes}"


    def getTaskType(self, tow):      # get task type from apn1225, 1328 and 1498
        return {
            "fot":"X",
            "sgh":"",
            "r/i":"",
            "ts":"X",
            "mod":"",
            "rep":"X" if tow == "AD_CMS" else "",
            "insp":"X",
        }

    def getCRS(self, row1225, row1328):
        serial = row1328[self.serial1328].value
        part = row1328[self.part1328].value
        sbt=row1225[self.sbt].value
        data = row1225[self.date].value
        for rows in self.ws1498:
            if rows[self.serial1498].value == serial:
                if rows[self.part1498].value == part:
                    if rows[self.date1498].value == data:
                        if rows[self.crs].value == sbt:
                            return "X"        
        return ""



    def save(self):     # save logbook 
        self.wblb.save("Logbook.xlsx")

    def getOperation(self, crs):
        if crs == "X":
            return "Maintenance, TBS, Repair, CRS"
        else:
            return "Maintenance, TBS, Repair"


    def getATA(self, pn):
        for rows in self.wsDatabase:
            if rows[0].value == pn:
                return rows[2].value

    
    def getDescription(self, pn):
        for rows in self.wsDatabase:
            if rows[0].value == pn:
                return rows[1]